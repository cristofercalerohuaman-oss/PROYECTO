# export_all_tables.py (versión segura y con fallback)
import os
from dotenv import load_dotenv
import pandas as pd
from sqlalchemy import create_engine, inspect, text
from openpyxl import load_workbook
from urllib.parse import quote_plus

load_dotenv()

OUTFILE = "base_completa.xlsx"

def engine_sqlite():
    sqlite_path = os.path.abspath("db.sqlite3")
    if os.path.exists(sqlite_path):
        print("Usando SQLite:", sqlite_path)
        return create_engine(f"sqlite:///{sqlite_path}")
    return None

def engine_postgres():
    DB_USER = os.getenv("DB_USER")
    DB_PASS = os.getenv("DB_PASS") or ""
    DB_HOST = os.getenv("DB_HOST", "localhost")
    DB_PORT = os.getenv("DB_PORT", "5432")
    DB_NAME = os.getenv("DB_NAME")

    if not DB_NAME:
        raise RuntimeError("No se encontró DB_NAME en .env. Si usas SQLite, crea db.sqlite3 o pon DB_NAME en .env")

    # percent-encode para evitar problemas con caracteres especiales
    user_enc = quote_plus(DB_USER) if DB_USER else ""
    pass_enc = quote_plus(DB_PASS)
    url = f"postgresql+psycopg2://{user_enc}:{pass_enc}@{DB_HOST}:{DB_PORT}/{DB_NAME}"

    # Intentar UTF8, luego LATIN1
    for enc in ("UTF8", "LATIN1"):
        try:
            print(f"Intentando conectar a Postgres con client_encoding={enc} ...")
            engine = create_engine(url, connect_args={"options": f"-c client_encoding={enc}"})
            # prueba rápida de conexión
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            print("Conexión OK con client_encoding=", enc)
            return engine
        except Exception as e:
            print(f"Fallo con client_encoding={enc}: {repr(e)}")
    # si falla todo:
    raise RuntimeError("No se pudo conectar a Postgres con UTF8/LATIN1. Revisa .env (password/encoding).")

def get_engine():
    # preferir sqlite si existe (evita psycopg2 y problemas de encoding)
    eng = engine_sqlite()
    if eng:
        return eng
    return engine_postgres()

def autofit_and_filter(path):
    wb = load_workbook(path)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    l = len(str(cell.value))
                    if l > max_len: max_len = l
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
    wb.save(path)

def normalizar_df(df):
    # intenta mantener texto legible: si hay bytes raros, reemplaza con �
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].apply(lambda x: x if not isinstance(x, str) else x)
    return df

def main():
    try:
        engine = get_engine()
    except Exception as e:
        print("ERROR al crear engine:", repr(e))
        print("Sugerencias: ")
        print(" - Si usas SQLite, coloca db.sqlite3 en la carpeta del proyecto.")
        print(" - Si usas Postgres, revisa tu archivo .env y usa una contraseña sólo con caracteres ASCII para probar.")
        print(" - Puedes probar: set PGCLIENTENCODING=LATIN1  (en Windows CMD) antes de ejecutar.")
        return

    insp = inspect(engine)
    try:
        # si es sqlite no hay schema='public'
        try:
            tables = insp.get_table_names(schema='public')
            if not tables:
                # sqlite: get_table_names may return empty for schema public -> fallback
                tables = insp.get_table_names()
        except Exception:
            tables = insp.get_table_names()
    except Exception as e:
        print("ERROR inspeccionando engine:", repr(e))
        return

    print("Tablas detectadas:", tables)
    if not tables:
        print("No se detectaron tablas. Salir.")
        return

    with pd.ExcelWriter(OUTFILE, engine="openpyxl") as writer:
        for t in tables:
            print("Exportando:", t)
            try:
                df = pd.read_sql(f"SELECT * FROM {t};", engine)
                df = normalizar_df(df)
                df.to_excel(writer, sheet_name=(t[:30]), index=False)
            except Exception as e:
                print(f"ERROR exportando tabla {t}: {repr(e)}")
    autofit_and_filter(OUTFILE)
    print("✅ Export completo a:", OUTFILE)

if __name__ == "__main__":
    main()
